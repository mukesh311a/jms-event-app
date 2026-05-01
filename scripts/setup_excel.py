"""
Build JMS-Event-Data.xlsx from CSV files in ../data/
Run: python scripts/setup_excel.py
"""
from pathlib import Path
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "JMS-Event-Data.xlsx"
DATA_DIR = ROOT / "data"

# Sheet order: participants first, then admin helpers, then masters
TABLES = [
    "Registrations",
    "FamilyMembers",
    "Donations",
    "CheckIn",
    "AdminPortalAttempts",
    "AdminConfig",
    "ProfessionMaster",
    "DesignationMaster",
]


def csv_rows(csv_path: Path):
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.reader(f))


def ensure_workbook(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.save(path)
    return load_workbook(path)


def clear_sheet(sheet):
    if sheet.max_row > 1:
        sheet.delete_rows(1, sheet.max_row)


def remove_existing_tables(sheet):
    if not sheet.tables:
        return
    for name in list(sheet.tables.keys()):
        del sheet.tables[name]


def write_rows(sheet, rows):
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)


def add_table(sheet, table_name: str):
    max_row = max(1, sheet.max_row)
    max_col = max(1, sheet.max_column)
    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(tab)


def autosize(sheet):
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def main():
    wb = ensure_workbook(XLSX_PATH)

    for table_name in TABLES:
        csv_path = DATA_DIR / f"{table_name}.csv"
        if not csv_path.exists():
            raise FileNotFoundError(f"Missing CSV: {csv_path}")

        rows = csv_rows(csv_path)
        if table_name in wb.sheetnames:
            ws = wb[table_name]
        else:
            ws = wb.create_sheet(table_name)

        remove_existing_tables(ws)
        clear_sheet(ws)
        write_rows(ws, rows)
        add_table(ws, table_name)
        autosize(ws)

    # Remove stray default sheets
    allowed = set(TABLES)
    for name in list(wb.sheetnames):
        if name not in allowed:
            wb.remove(wb[name])

    wb.save(XLSX_PATH)
    print(f"OK: {XLSX_PATH}")


if __name__ == "__main__":
    main()
